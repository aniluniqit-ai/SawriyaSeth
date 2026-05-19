"""Excel Trade Logger — LKS V10"""
import os, logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger("ExcelTracker")

HEADERS = [
    "Trade ID", "Date", "Time", "Channel", "Symbol", "Option Type",
    "Strike", "Entry", "Qty", "Lot", "SL", "Target 1", "Target 2",
    "Exit Price", "PnL", "Status", "Exit Reason", "Mode", "Signal Text"
]

EXCEL_DIR = "trades"

class ExcelTracker:
    def __init__(self):
        os.makedirs(EXCEL_DIR, exist_ok=True)
        self._file = None
        self._wb = None
        self._ws = None
        self._ensure_file()

    def _path(self):
        today = datetime.now().strftime("%Y%m%d")
        return os.path.join(EXCEL_DIR, f"Trades_Log_{today}.xlsx")

    def _ensure_file(self):
        path = self._path()
        if os.path.exists(path):
            try:
                self._wb = load_workbook(path)
                self._ws = self._wb.active
                self._file = path
                return
            except Exception:
                pass
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = "Trades"
        self._write_headers()
        self._file = path
        self._wb.save(self._file)
        logger.info(f"Excel file created: {self._file}")

    def _write_headers(self):
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
        thin = Side(style="thin", color="999999")
        for col, h in enumerate(HEADERS, 1):
            cell = self._ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        self._ws.column_dimensions["A"].width = 12
        self._ws.column_dimensions["B"].width = 12
        self._ws.column_dimensions["C"].width = 10
        self._ws.column_dimensions["D"].width = 28
        self._ws.column_dimensions["E"].width = 14
        self._ws.column_dimensions["F"].width = 12
        self._ws.column_dimensions["G"].width = 10
        self._ws.column_dimensions["H"].width = 10
        self._ws.column_dimensions["I"].width = 8
        self._ws.column_dimensions["J"].width = 6
        self._ws.column_dimensions["K"].width = 10
        self._ws.column_dimensions["L"].width = 10
        self._ws.column_dimensions["M"].width = 10
        self._ws.column_dimensions["N"].width = 10
        self._ws.column_dimensions["O"].width = 10
        self._ws.column_dimensions["P"].width = 12
        self._ws.column_dimensions["Q"].width = 18
        self._ws.column_dimensions["R"].width = 10
        self._ws.column_dimensions["S"].width = 40

    def _next_row(self):
        return self._ws.max_row + 1

    def log_signal(self, signal):
        """Log a parsed signal (even if rejected)"""
        row = self._next_row()
        now = datetime.now()
        data = {
            "A": str(now.strftime("%Y%m%d")),
            "B": now.strftime("%Y-%m-%d"),
            "C": now.strftime("%H:%M:%S"),
            "D": getattr(signal, "channel", ""),
            "E": getattr(signal, "symbol", ""),
            "F": getattr(signal, "option_type", ""),
            "G": getattr(signal, "strike", ""),
            "H": getattr(signal, "entry", ""),
            "P": "SIGNAL",
            "Q": "Parsed",
            "S": (getattr(signal, "raw", "") or "")[:100],
        }
        for col, val in data.items():
            self._ws[f"{col}{row}"] = val
        self._save()

    def log_trade_open(self, trade_data: dict):
        """Log an opened position"""
        row = self._next_row()
        now = datetime.now()
        pnl_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        data = {
            "A": trade_data.get("id", ""),
            "B": now.strftime("%Y-%m-%d"),
            "C": now.strftime("%H:%M:%S"),
            "D": trade_data.get("source", ""),
            "E": trade_data.get("segment", ""),
            "F": trade_data.get("option_type", ""),
            "G": trade_data.get("strike", ""),
            "H": trade_data.get("entry", ""),
            "I": trade_data.get("qty", ""),
            "J": trade_data.get("lot", ""),
            "K": trade_data.get("sl", ""),
            "L": (trade_data.get("targets") or [None, None])[0] if isinstance(trade_data.get("targets"), list) else "",
            "M": (trade_data.get("targets") or [None, None])[1] if isinstance(trade_data.get("targets"), list) and len(trade_data.get("targets", [])) > 1 else "",
            "P": "OPEN",
            "R": trade_data.get("mode", "paper"),
        }
        for col, val in data.items():
            cell = self._ws[f"{col}{row}"]
            cell.value = val
            if val and col in ("P",):
                cell.fill = pnl_fill
        self._save()

    def log_trade_close(self, close_data: dict):
        """Update the trade row with exit details, or append if not found"""
        trade_id = close_data.get("id", "")
        found = False
        for row in range(2, self._ws.max_row + 1):
            if self._ws[f"A{row}"].value == trade_id:
                now = datetime.now()
                updates = {
                    "C": now.strftime("%H:%M:%S"),
                    "N": close_data.get("exit_price", ""),
                    "O": close_data.get("pnl", ""),
                    "P": "CLOSED",
                    "Q": close_data.get("reason", ""),
                }
                for col, val in updates.items():
                    cell = self._ws[f"{col}{row}"]
                    cell.value = val
                    if col == "O":
                        pnl = float(val) if val else 0
                        if pnl > 0:
                            cell.fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
                            cell.font = Font(color="006100", bold=True)
                        elif pnl < 0:
                            cell.fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
                            cell.font = Font(color="9c0006", bold=True)
                found = True
                break
        if not found:
            row = self._next_row()
            now = datetime.now()
            data = {
                "A": trade_id,
                "B": now.strftime("%Y-%m-%d"),
                "C": now.strftime("%H:%M:%S"),
                "D": close_data.get("source", ""),
                "E": close_data.get("segment", ""),
                "F": close_data.get("option_type", ""),
                "G": close_data.get("strike", ""),
                "H": close_data.get("entry", ""),
                "I": close_data.get("qty", ""),
                "K": close_data.get("sl", ""),
                "N": close_data.get("exit_price", ""),
                "O": close_data.get("pnl", ""),
                "P": "CLOSED",
                "Q": close_data.get("reason", ""),
            }
            for col, val in data.items():
                self._ws[f"{col}{row}"] = val
        self._save()

    def _save(self):
        try:
            self._wb.save(self._file)
        except Exception as e:
            logger.warning(f"Excel save error: {e}")
